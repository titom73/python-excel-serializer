import os
from optparse import OptionParser
import pprint
import sys
import json
import openpyxl
import itertools

class ExcelSerializer:
	""" Convert Excel data into python structures.

	Allow to extract data from an Excel sheet to convert it to different 
	This class supports import from different tab from Excel file and can serialize both: table and list

	Output example for a table:
	table = ExcelSerializer(excel_path='../excel/input.xlsx')
	table.serialized_table(sheet="Topology", nb_columns=6)
	print table.get_data(sheet="Topology")

	[   {   'id': '1',
		u'local_device': u'dev1',
		u'local_port': u'port1',
		u'local_port_name': u'et-0/0/0',
		u'remote_device': u'dev2',
		u'remote_port': u'port1',
		u'remote_port_name': u'et-0/0/0'},
	{   'id': '2',
		u'local_device': u'dev1',
		u'local_port': u'port2',
		u'local_port_name': u'et-0/0/1',
		u'remote_device': u'dev3',
		u'remote_port': u'port1',
		u'remote_port_name': u'et-0/0/1'},
	{   'id': '3',
		u'local_device': u'dev2',
		u'local_port': u'port3',
		u'local_port_name': u'et-0/0/2',
		u'remote_device': u'dev4',
		u'remote_port': u'port1',
		u'remote_port_name': u'et-0/0/2'}]

	Output example for a list:
	table = ExcelSerializer(excel_path='../excel/input.xlsx')
	table.serialized_list(sheet="List")
	print table.get_data(sheet="List")
	{   'asn_base': '65000',
		'bgp_export': 'underlay-export',
		'bgp_group': 'underlay',
		'bgp_import': 'underlay-import',
		'mtu_phy_int': ['9200', '1000']
	}

	"""
	def __init__(self, excel_path = None, sheet="Sheet1", has_header=True, nb_columns=0):
		""" Class constructor

		Keyword arguments:
		excel_path -- Path to excel file you want to read data.
		sheet -- Tab name where data are located. (Default is Sheet1)
		has_header -- Boolean to enable header finding.
		nb_columns -- Number of columns to look for data

		Return Value:
		N/A

		"""
		self._yaml=dict()
		self._filename = excel_path

	def serialize_table(self, sheet=None, nb_columns=2, has_header=True):
		""" Serialize Excel file into python structure
	
		Open Excel file defined in self._filename and look for data
		Because columns are using Letters instead of numbers, it is required to convert 
		Decimal value to alphabet letter.

		For every line, function is reading every cell and move them to a dict entry
		Once line is completely read, dict is appended to a list of all lines.

		Complete value dict is saved directly in self._data_array
		
		Keyword arguments:
		none -- none

		Return Value:
		N/A

		"""
		wb = openpyxl.load_workbook(self._filename)
		ws = wb[sheet]
		_data_array=[]
		_header = self._find_header(sheet=sheet, nb_columns=nb_columns)
		# because Excel columns are alphabet letters instead of number,
		# we have to convert decimal to letter. 
		# First letter 'a' is identify with code 97
		# Max column is 97(a) + nb_columns -1
		col_max = chr(97-1+nb_columns)
		cell_range = 'a{}:{}{}'.format(ws.min_row+1 , col_max, ws.max_row)
		row_id = 1
		# Extract ROW from the range
		for row in ws[cell_range]:
			line = dict()
			cell_index = 0
			# Read every CELL part of the row and within our scope
			for cell in row:
				# Force casting to str for all cells. Useful if excel encode a value to int or boolean
				if has_header is True:
					line[_header[cell_index]] = str(cell.value).lower()
				else:
					line[cell_index] = str(cell.value).lower()
				cell_index +=1 
			# Add ROW_ID to create unique ID
			line['id'] = str(row_id)
			# Add line to the complete dict
			_data_array.append(line)
			row_id += 1
		self._yaml[sheet] = _data_array
		return self._yaml[sheet]

	def serialize_list(self, sheet=None, nb_columns=2):
		wb = openpyxl.load_workbook(self._filename)
		ws = wb[sheet]
		_data = dict()
		col_max = chr(97-1+nb_columns)
		cell_range = 'a{}:{}{}'.format(ws.min_row , col_max, ws.max_row)
		print cell_range
		for row in ws[cell_range]:
			_data[str(row[0].value).lower()] = self._serialize_cell( xlsCellStr = str(row[1].value).lower() )
		self._yaml[sheet] = _data
		return self._yaml[sheet]

	def _find_header(self, sheet=None, nb_columns=2):
		""" Look for array in the table and provide list of them
	
		Extract table header and create a dict() to store them with a standardize name: lower and replace space by "_"
		
		Keyword arguments:
		none -- none

		Return Value:
		N/A

		"""
		header = []
		wb = openpyxl.load_workbook(self._filename)
		ws = wb[sheet]
		col_max = chr(97-1+nb_columns)
		cell_range = 'a{}:{}{}'.format(ws.min_row, col_max, 1)
		for row in ws[cell_range]:
			for cell in row:
				header.append( self._string_cleanup(cell.value) )
		return header

	def _serialize_cell( self, xlsCellStr=None):
		print "Length: "+str(len(str(xlsCellStr).split('\n')))
		if len(str(xlsCellStr).split('\n')) > 1:
			return str(xlsCellStr).split('\n')
		else:
			return xlsCellStr

	def _string_cleanup( self, string ):
		""" Cleanup a string to lower and remove space

		Keyword arguments:
		string -- string to cleanup

		Return Value:
		cleanup string

		"""
		return string.replace(' ', '_').lower()

	def get_data(self, sheet=None):
		""" Provide a read only access to table data

		Keyword arguments:
		none -- none

		Return Value:
		self._array_data[]

		"""
		return self.get_yaml(sheet=sheet)

	def get_header(self):
		""" Provide a read only access to header table data

		Keyword arguments:
		none -- none

		Return Value:
		self._header[]

		"""
		return self._header

	def get_yaml(self, sheet=None):
		""" Return a per sheet Puthon structure compatible with YAML language

		Keyword arguments:
		none -- none

		Return Value:
		dict[self._header] = self._array_data[]

		"""
		if sheet in self._yaml:
			return self._yaml[sheet]
		else:
			return None
	def get_yaml_all(self):
		return self._yaml